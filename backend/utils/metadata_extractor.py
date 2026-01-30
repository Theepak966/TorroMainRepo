import hashlib
import json
import logging
from typing import Dict, Optional, List
import pyarrow.parquet as pq
import io
import csv
from collections import Counter
import xml.etree.ElementTree as ET

logger = logging.getLogger(__name__)


try:
    import fastavro
    AVRO_AVAILABLE = True
except ImportError:
    AVRO_AVAILABLE = False
    logger.warning('FN:metadata_extractor_import AVRO_AVAILABLE:{}'.format(False))

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning('FN:metadata_extractor_import EXCEL_AVAILABLE:{}'.format(False))

try:
    import xmltodict
    XML_AVAILABLE = True
except ImportError:
    XML_AVAILABLE = False
    logger.warning('FN:metadata_extractor_import XML_AVAILABLE:{}'.format(False))

logger = logging.getLogger(__name__)


try:
    # Azure DLP is intentionally disabled for performance; keep only the pattern-based detector entrypoint.
    from .azure_dlp_client import detect_pii_in_column
    AZURE_DLP_AVAILABLE = False
    logger.info('FN:metadata_extractor_import AZURE_DLP_AVAILABLE:{} message:Azure DLP disabled; using pattern-based detection'.format(False))
except ImportError:
    AZURE_DLP_AVAILABLE = False
    logger.warning('FN:metadata_extractor_import AZURE_DLP_AVAILABLE:{} message:PII detector module not available'.format(False))

    def detect_pii_in_column(column_name: str, sample_data=None) -> Dict:
        return {"pii_detected": False, "pii_types": []}


def generate_file_hash(file_content: bytes) -> str:

    hash_obj = hashlib.shake_128(file_content)
    return hash_obj.hexdigest(16)


def extract_parquet_schema(file_content: bytes, include_pii_detection: bool = False, sample_data: Optional[Dict] = None) -> Dict:
    """
    Extract schema from parquet file content.
    Robust implementation that handles various parquet file formats and edge cases.
    
    Args:
        file_content: Parquet file bytes (can be just footer for schema, or full file for PII)
        include_pii_detection: If True, attempts to extract sample data for PII detection
        sample_data: Optional pre-extracted sample data (for when footer-only is used)
    
    Returns:
        Dictionary with columns, types, and optionally PII detection
    """
    if not file_content or len(file_content) == 0:
        logger.warning('FN:extract_parquet_schema message:Empty file content provided')
        return {"columns": [], "num_columns": 0}
    
    try:
        # Validate parquet file format
        if len(file_content) < 8 or file_content[-4:] != b'PAR1':
            logger.warning('FN:extract_parquet_schema file_content_size:{} message:Invalid parquet file (missing PAR1 magic)'.format(len(file_content)))
            return {"columns": [], "num_columns": 0}
        
        # Create ParquetFile object - this validates the file structure
        # Note: For combined structures (row group + footer), PyArrow might fail to parse
        # but it should still be able to extract schema from the footer portion
        try:
            parquet_file = pq.ParquetFile(io.BytesIO(file_content))
        except (IOError, OSError, ValueError, TypeError) as e:
            # If combined structure fails, try to extract schema from footer only
            # Footer should be at the end of file_content (has PAR1 magic)
            if len(file_content) > 1000 and file_content[-4:] == b'PAR1':
                logger.debug('FN:extract_parquet_schema message:Combined structure failed, attempting footer-only extraction error:{}'.format(str(e)))
                try:
                    # Prefer an exact footer slice using the Parquet footer length stored in the last 8 bytes:
                    # [4 bytes footer_len (LE)] + b'PAR1'
                    # This is critical for very wide schemas where footer can exceed 2MB.
                    import struct
                    footer_only = None
                    try:
                        footer_len = struct.unpack("<I", file_content[-8:-4])[0]
                        # Sanity bounds: cap at 64MB to avoid pathological memory use while supporting
                        # very large footers (e.g. many row-groups + wide schemas).
                        max_footer = 64 * 1024 * 1024
                        if 0 < footer_len <= max_footer:
                            start = max(0, len(file_content) - (footer_len + 8))
                            footer_only = file_content[start:]
                    except Exception:
                        footer_only = None

                    # Fallback: take the last 8MB (older behavior was 2MB, which is too small for wide schemas)
                    if footer_only is None:
                        footer_start = max(0, len(file_content) - 8 * 1024 * 1024)
                        footer_only = file_content[footer_start:]

                    if len(footer_only) >= 8 and footer_only[-4:] == b'PAR1':
                        parquet_file = pq.ParquetFile(io.BytesIO(footer_only))
                        logger.info('FN:extract_parquet_schema message:Successfully extracted schema from footer-only after combined structure failed')
                    else:
                        logger.warning('FN:extract_parquet_schema invalid_parquet_file error:{}'.format(str(e)))
                        return {"columns": [], "num_columns": 0}
                except Exception as footer_error:
                    logger.warning('FN:extract_parquet_schema invalid_parquet_file error:{} footer_fallback_error:{}'.format(str(e), str(footer_error)))
                    return {"columns": [], "num_columns": 0}
            else:
                logger.warning('FN:extract_parquet_schema invalid_parquet_file error:{}'.format(str(e)))
                return {"columns": [], "num_columns": 0}
        except Exception as e:
            logger.error('FN:extract_parquet_schema unexpected_error error:{}'.format(str(e)), exc_info=True)
            return {"columns": [], "num_columns": 0}
        
        # Get schema
        try:
            schema = parquet_file.schema_arrow
        except Exception as e:
            logger.warning('FN:extract_parquet_schema message:Could not read schema error:{}'.format(str(e)))
            return {"columns": [], "num_columns": 0}
        
        if schema is None or len(schema) == 0:
            logger.warning('FN:extract_parquet_schema message:Empty schema found')
            return {"columns": [], "num_columns": 0}
        
        # Try to get sample data for PII detection if requested
        # For PII detection to work properly, we need actual data samples, not just column names
        # This works with full file downloads or downloads that include row group data
        if include_pii_detection and sample_data is None:
            try:
                # Check if we have enough data to read row groups
                # If file content is > 50KB, we likely have row group data (not just footer)
                has_row_group_data = len(file_content) > 50000
                
                if has_row_group_data and parquet_file.num_row_groups > 0:
                    try:
                        # Try to read first row group for sample data
                        table = parquet_file.read_row_group(0)
                        if table is not None and len(table) > 0:
                            # Get up to 20 rows for better PII detection accuracy
                            sample_size = min(20, len(table))
                            sample_table = table.slice(0, sample_size)
                            sample_data = sample_table.to_pandas().to_dict('list')
                            logger.info('FN:extract_parquet_schema message:Extracted {} rows from row group for PII detection'.format(sample_size))
                        else:
                            logger.debug('FN:extract_parquet_schema message:Row group exists but is empty')
                            sample_data = None
                    except Exception as e:
                        logger.debug('FN:extract_parquet_schema message:Could not read row group data error:{}'.format(str(e)))
                        sample_data = None
                else:
                    # Footer-only mode: no row group data available
                    # PII detection will use column names only (less accurate but still works)
                    logger.debug('FN:extract_parquet_schema message:Footer-only mode, PII detection will use column names only')
                    sample_data = None
            except Exception as e:
                logger.debug('FN:extract_parquet_schema message:Error extracting sample data error:{}'.format(str(e)))
                sample_data = None
        
        columns = []
        for i in range(len(schema)):
            try:
                field = schema.field(i)
                
                if field is None:
                    logger.warning('FN:extract_parquet_schema field_index:{} message:Null field found, skipping'.format(i))
                    continue
                
                # Get column samples for PII detection
                # PII detection works best with actual data samples, but falls back to column names
                column_samples = None
                if include_pii_detection:
                    if sample_data and field.name in sample_data:
                        # Extract up to 20 sample values for better PII detection accuracy
                        column_samples = [str(val) for val in sample_data[field.name][:20] if val is not None]
                        logger.debug('FN:extract_parquet_schema field:{} samples:{} message:Using {} data samples for PII detection'.format(
                            field.name, len(column_samples), len(column_samples)
                        ))
                    else:
                        # Fallback to column name only (no data samples available - footer-only mode)
                        column_samples = None
                        logger.debug('FN:extract_parquet_schema field:{} message:Using column name only for PII detection (no data samples)'.format(field.name))
                
                # PII detection: Use column name + samples if available (best accuracy), or column name only (fallback)
                pii_result = detect_pii_in_column(field.name, column_samples)
                
                # Extract type information with better handling of complex types
                arrow_type = field.type
                type_str = str(arrow_type)
                
                # Enhanced type information for complex types
                type_info = {
                    "base_type": type_str,
                    "is_nested": False,
                    "is_nullable": field.nullable
                }
                
                # Handle nested/complex types
                try:
                    import pyarrow as pa
                    if isinstance(arrow_type, pa.StructType):
                        type_info["is_nested"] = True
                        type_info["nested_type"] = "struct"
                        type_info["fields"] = {f.name: str(f.type) for f in arrow_type}
                    elif isinstance(arrow_type, pa.ListType):
                        type_info["is_nested"] = True
                        type_info["nested_type"] = "list"
                        type_info["element_type"] = str(arrow_type.value_type)
                    elif isinstance(arrow_type, pa.MapType):
                        type_info["is_nested"] = True
                        type_info["nested_type"] = "map"
                        type_info["key_type"] = str(arrow_type.key_type)
                        type_info["value_type"] = str(arrow_type.value_type)
                except Exception as e:
                    logger.debug('FN:extract_parquet_schema field:{} message:Could not extract nested type info error:{}'.format(field.name, str(e)))
                
                column_data = {
                    "name": field.name or f"column_{i}",  # Fallback if name is None
                    "type": type_str,
                    "type_info": type_info,
                    "nullable": field.nullable
                }
                
                # Add PII detection results
                if include_pii_detection:
                    if pii_result.get("pii_detected"):
                        column_data["pii_detected"] = True
                        column_data["pii_types"] = pii_result.get("pii_types", [])
                    else:
                        column_data["pii_detected"] = False
                        column_data["pii_types"] = None
                else:
                    # PII detection skipped - mark as None
                    column_data["pii_detected"] = None
                    column_data["pii_types"] = None
                
                columns.append(column_data)
                
            except Exception as e:
                logger.warning('FN:extract_parquet_schema field_index:{} error:{} message:Skipping field due to error'.format(i, str(e)))
                continue
        
        if len(columns) == 0:
            logger.warning('FN:extract_parquet_schema message:No columns extracted from schema')
            return {"columns": [], "num_columns": 0}
        
        schema_dict = {
            "columns": columns,
            "num_columns": len(columns),
        }
        
        # Extract additional metadata if available
        try:
            metadata = parquet_file.metadata
            if metadata:
                if hasattr(metadata, 'num_rows') and metadata.num_rows is not None:
                    schema_dict["num_rows"] = metadata.num_rows
                
                if hasattr(metadata, 'num_row_groups'):
                    schema_dict["num_row_groups"] = metadata.num_row_groups
                
                # Extract file-level metadata if available
                if hasattr(metadata, 'metadata') and metadata.metadata:
                    try:
                        # Parquet metadata is stored as bytes, try to decode
                        file_metadata = {}
                        for key, value in metadata.metadata.items():
                            try:
                                if isinstance(key, bytes):
                                    key = key.decode('utf-8', errors='ignore')
                                if isinstance(value, bytes):
                                    value = value.decode('utf-8', errors='ignore')
                                file_metadata[key] = value
                            except Exception:
                                pass
                        if file_metadata:
                            schema_dict["file_metadata"] = file_metadata
                    except Exception as e:
                        logger.debug('FN:extract_parquet_schema message:Could not extract file metadata error:{}'.format(str(e)))
        except Exception as e:
            logger.debug('FN:extract_parquet_schema message:Could not extract metadata error:{}'.format(str(e)))
        
        logger.info('FN:extract_parquet_schema num_columns:{} num_rows:{}'.format(
            len(columns), schema_dict.get("num_rows", "unknown")
        ))
        
        return schema_dict
        
    except Exception as e:
        logger.error('FN:extract_parquet_schema file_content_size:{} error:{}'.format(
            len(file_content) if file_content else 0, str(e)
        ), exc_info=True)
        return {"columns": [], "num_columns": 0}


def generate_schema_hash(schema_json: Dict) -> str:
    schema_str = json.dumps(schema_json, sort_keys=True)
    hash_obj = hashlib.shake_128(schema_str.encode())
    return hash_obj.hexdigest(16)


def extract_csv_schema(file_content: bytes, sample_size: int = 0) -> Dict:

    try:
        content_str = file_content.decode('utf-8', errors='ignore')
        lines = content_str.split('\n')
        
        if not lines:
            return {"columns": [], "num_columns": 0}
        

        reader = csv.reader([lines[0]])
        headers = next(reader, None)
        
        if not headers:
            return {"columns": [], "num_columns": 0}
        

        sample_rows = []
        for line in lines[1:11]:
            if line.strip():
                try:
                    row_reader = csv.reader([line])
                    row = next(row_reader, None)
                    if row and len(row) == len(headers):
                        sample_rows.append(row)
                except Exception:
                    continue
        
        columns = []
        
        for i, header in enumerate(headers):
            header = header.strip()
            if not header:
                header = f"column_{i+1}"
            

            column_samples = []
            if sample_rows:
                for row in sample_rows:
                    if i < len(row) and row[i]:
                        column_samples.append(str(row[i]).strip())
            

            pii_result = detect_pii_in_column(header, column_samples if column_samples else None)
            
            column_data = {
                "name": header,
                "type": "string",
                "nullable": True
            }
            

            if pii_result.get("pii_detected"):
                column_data["pii_detected"] = True
                column_data["pii_types"] = pii_result.get("pii_types", [])
            else:
                column_data["pii_detected"] = False
                column_data["pii_types"] = None
            
            columns.append(column_data)
        
        return {
            "columns": columns,
            "num_columns": len(columns),
            "num_rows": None,
            "has_header": True,
            "delimiter": ","
        }
        
    except Exception as e:
        logger.warning('FN:extract_csv_schema file_content_size:{} sample_size:{} error:{}'.format(len(file_content) if file_content else 0, sample_size, str(e)))
        return {"columns": [], "num_columns": 0}


def infer_column_type(values: List[str]) -> str:
    if not values:
        return "string"
    
    non_empty = [v.strip() for v in values if v and v.strip()]
    if not non_empty:
        return "string"
    
    int_count = 0
    float_count = 0
    bool_count = 0
    date_count = 0
    
    for value in non_empty[:100]:
        value = value.strip()
        
        if value.lower() in ['true', 'false', 'yes', 'no', '1', '0']:
            bool_count += 1
        
        try:
            int(value)
            int_count += 1
        except ValueError:
            pass
        
        try:
            float(value)
            float_count += 1
        except ValueError:
            pass
        
        if '/' in value or '-' in value:
            parts = value.replace('/', '-').split('-')
            if len(parts) == 3:
                try:
                    int(parts[0])
                    int(parts[1])
                    int(parts[2])
                    date_count += 1
                except ValueError:
                    pass
    
    total = len(non_empty)
    threshold = total * 0.8
    
    if int_count >= threshold:
        return "int64"
    elif float_count >= threshold:
        return "double"
    elif bool_count >= threshold:
        return "bool"
    elif date_count >= threshold:
        return "date"
    else:
        return "string"


def extract_json_schema(file_content: bytes) -> Dict:
    try:
        content_str = file_content.decode('utf-8', errors='ignore').strip()
        
        if not content_str:
            return {"columns": [], "num_columns": 0}
        
        try:
            data = json.loads(content_str)
        except json.JSONDecodeError:
            return {"columns": [], "num_columns": 0, "error": "Invalid JSON"}
        
        columns = []
        
        if isinstance(data, dict):

            for key in data.keys():

                sample_value = str(data.get(key, "")) if data.get(key) is not None else None
                column_samples = [sample_value] if sample_value else None
                

                pii_result = detect_pii_in_column(str(key), column_samples)
                
                column_data = {
                    "name": str(key),
                    "type": "string",
                    "nullable": True
                }
                

                # Use PII detection result (Azure DLP with pattern-based fallback)
                if pii_result.get("pii_detected"):
                    column_data["pii_detected"] = True
                    column_data["pii_types"] = pii_result.get("pii_types", [])
                else:
                    column_data["pii_detected"] = False
                    column_data["pii_types"] = None
                
                columns.append(column_data)
        elif isinstance(data, list) and len(data) > 0:
            first_item = data[0]
            if isinstance(first_item, dict):

                for key in first_item.keys():

                    column_samples = []
                    for item in data[:10]:
                        if isinstance(item, dict) and key in item and item[key] is not None:
                            column_samples.append(str(item[key]))
                    

                    pii_result = detect_pii_in_column(str(key), column_samples if column_samples else None)
                    
                    column_data = {
                        "name": str(key),
                        "type": "string",
                        "nullable": True
                    }
                    

                    if pii_result.get("pii_detected"):
                        column_data["pii_detected"] = True
                        column_data["pii_types"] = pii_result.get("pii_types", [])
                    else:
                        column_data["pii_detected"] = False
                        column_data["pii_types"] = None
                    
                    columns.append(column_data)
            else:
                columns.append({
                    "name": "value",
                    "type": "string",
                    "nullable": True,
                    "pii_detected": False,
                    "pii_types": None
                })
        
        return {
            "columns": columns,
            "num_columns": len(columns),
            "structure": "object" if isinstance(data, dict) else "array",
            "num_items": len(data) if isinstance(data, list) else None
        }
        
    except Exception as e:
        logger.warning('FN:extract_json_schema file_content_size:{} error:{}'.format(len(file_content) if file_content else 0, str(e)))
        return {"columns": [], "num_columns": 0}


def infer_json_type(value) -> str:
    if value is None:
        return "null"
    elif isinstance(value, bool):
        return "bool"
    elif isinstance(value, int):
        return "int64"
    elif isinstance(value, float):
        return "double"
    elif isinstance(value, str):
        return "string"
    elif isinstance(value, list):
        if len(value) > 0:
            return f"array<{infer_json_type(value[0])}>"
        return "array"
    elif isinstance(value, dict):
        return "object"
    else:
        return "string"


def extract_avro_schema(file_content: bytes) -> Dict:
    if not AVRO_AVAILABLE:
        return {"columns": [], "num_columns": 0, "error": "Avro library not available"}
    
    try:
        file_obj = io.BytesIO(file_content)
        reader = fastavro.reader(file_obj)
        schema = reader.schema
        
        columns = []
        if isinstance(schema, dict) and 'fields' in schema:
            for field in schema['fields']:
                field_name = field.get('name', 'unknown')
                field_type = str(field.get('type', 'unknown'))
                

                pii_result = detect_pii_in_column(field_name, None)
                
                column_data = {
                    "name": field_name,
                    "type": field_type,
                    "nullable": True
                }
                
                # Use PII detection result (Azure DLP with pattern-based fallback)
                if pii_result.get("pii_detected"):
                    column_data["pii_detected"] = True
                    column_data["pii_types"] = pii_result.get("pii_types", [])
                else:
                    column_data["pii_detected"] = False
                    column_data["pii_types"] = None
                
                columns.append(column_data)
        
        return {
            "columns": columns,
            "num_columns": len(columns),
            "avro_schema": schema
        }
    except Exception as e:
        logger.warning('FN:extract_avro_schema error:{}'.format(str(e)))
        return {"columns": [], "num_columns": 0}


def extract_excel_schema(file_content: bytes) -> Dict:
    if not EXCEL_AVAILABLE:
        return {"columns": [], "num_columns": 0, "error": "Excel library not available"}
    
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
        sheet = workbook.active
        
        if not sheet:
            return {"columns": [], "num_columns": 0}
        

        headers = []
        for cell in sheet[1]:
            headers.append(str(cell.value) if cell.value else f"column_{len(headers)+1}")
        

        sample_rows = []
        for row_idx in range(2, min(12, sheet.max_row + 1)):
            row_data = []
            for cell in sheet[row_idx]:
                row_data.append(str(cell.value) if cell.value else "")
            if any(row_data):
                sample_rows.append(row_data)
        
        columns = []
        for i, header in enumerate(headers):
            header = header.strip()
            if not header:
                header = f"column_{i+1}"
            

            column_samples = []
            for row in sample_rows:
                if i < len(row) and row[i]:
                    column_samples.append(str(row[i]).strip())
            

            pii_result = detect_pii_in_column(header, column_samples if column_samples else None)
            
            column_data = {
                "name": header,
                "type": "string",
                "nullable": True
            }
            
            if pii_result.get("pii_detected"):
                column_data["pii_detected"] = True
                column_data["pii_types"] = pii_result.get("pii_types", [])
            else:
                column_data["pii_detected"] = False
                column_data["pii_types"] = None
            
            columns.append(column_data)
        
        return {
            "columns": columns,
            "num_columns": len(columns),
            "num_rows": sheet.max_row - 1,
            "sheet_name": sheet.title
        }
    except Exception as e:
        logger.warning('FN:extract_excel_schema error:{}'.format(str(e)))
        return {"columns": [], "num_columns": 0}


def extract_xml_schema(file_content: bytes) -> Dict:
    try:
        content_str = file_content.decode('utf-8', errors='ignore')
        root = ET.fromstring(content_str)
        
        columns = []
        seen_tags = set()
        
        def extract_elements(element, path=""):
            current_path = f"{path}/{element.tag}" if path else element.tag
            
            if current_path not in seen_tags:
                seen_tags.add(current_path)
                

                text_value = element.text.strip() if element.text and element.text.strip() else None
                column_samples = [text_value] if text_value else None
                

                pii_result = detect_pii_in_column(element.tag, column_samples)
                
                column_data = {
                    "name": current_path,
                    "type": "string",
                    "nullable": True
                }
                
                # Use PII detection result (Azure DLP with pattern-based fallback)
                if pii_result.get("pii_detected"):
                    column_data["pii_detected"] = True
                    column_data["pii_types"] = pii_result.get("pii_types", [])
                else:
                    column_data["pii_detected"] = False
                    column_data["pii_types"] = None
                
                columns.append(column_data)
            

            for child in element:
                extract_elements(child, current_path)
        
        extract_elements(root)
        
        return {
            "columns": columns,
            "num_columns": len(columns),
            "root_element": root.tag
        }
    except Exception as e:
        logger.warning('FN:extract_xml_schema error:{}'.format(str(e)))
        return {"columns": [], "num_columns": 0}


def extract_orc_schema(file_content: bytes) -> Dict:
    try:

        import pyarrow.orc as orc
        orc_file = orc.ORCFile(io.BytesIO(file_content))
        schema = orc_file.schema
        
        columns = []
        for i in range(len(schema)):
            field = schema.field(i)
            

            pii_result = detect_pii_in_column(field.name, None)
            
            column_data = {
                "name": field.name,
                "type": str(field.type),
                "nullable": field.nullable
            }
            
            if pii_result.get("pii_detected"):
                column_data["pii_detected"] = True
                column_data["pii_types"] = pii_result.get("pii_types", [])
            else:
                column_data["pii_detected"] = False
                column_data["pii_types"] = None
            
            columns.append(column_data)
        
        return {
            "columns": columns,
            "num_columns": len(columns)
        }
    except ImportError:
        logger.warning('FN:extract_orc_schema message:PyArrow ORC support not available')
        return {"columns": [], "num_columns": 0, "error": "ORC support requires pyarrow with ORC support"}
    except Exception as e:
        logger.warning('FN:extract_orc_schema error:{}'.format(str(e)))
        return {"columns": [], "num_columns": 0}


def extract_delta_lake_schema(file_content: bytes, blob_path: str) -> Dict:


    try:

        parquet_schema = extract_parquet_schema(file_content)
        if parquet_schema.get("num_columns", 0) > 0:
            parquet_schema["format"] = "delta_lake"
            return parquet_schema
        else:
            return {"columns": [], "num_columns": 0, "format": "delta_lake"}
    except Exception as e:
        logger.warning('FN:extract_delta_lake_schema error:{}'.format(str(e)))
        return {"columns": [], "num_columns": 0}


def extract_file_metadata(blob_info: Dict, file_content: Optional[bytes] = None) -> Dict:
    from datetime import datetime
    
    file_name = blob_info["name"]
    file_extension = "." + file_name.split(".")[-1] if "." in file_name else ""
    file_format = file_extension[1:].lower() if file_extension else "unknown"
    


    is_data_lake = False
    if "/" in file_name or file_name.startswith("raw/") or file_name.startswith("processed/"):
        is_data_lake = True

        if not file_extension:

            if "parquet" in file_name.lower():
                file_format = "parquet"
            elif "csv" in file_name.lower():
                file_format = "csv"
            elif "json" in file_name.lower():
                file_format = "json"
            elif "avro" in file_name.lower():
                file_format = "avro"
            elif "orc" in file_name.lower():
                file_format = "orc"
            elif "delta" in file_name.lower() or "_delta_log" in file_name:
                file_format = "delta_lake"
    
    file_hash = None
    if file_content:
        file_hash = generate_file_hash(file_content)
    else:
        file_hash = generate_file_hash(b"")
    
    file_metadata = {
        "basic": {
            "name": file_name,
            "extension": file_extension,
            "format": file_format,
            "size_bytes": blob_info["size"],
            "content_type": blob_info.get("content_type", "application/octet-stream"),
            "mime_type": blob_info.get("content_type", "application/octet-stream")
        },
        "hash": {
            "algorithm": "shake128",
            "value": file_hash,
            "computed_at": datetime.utcnow().isoformat() + "Z"
        },
        "timestamps": {
            "created_at": blob_info["created_at"].isoformat() if blob_info.get("created_at") else None,
            "last_modified": blob_info["last_modified"].isoformat() if blob_info.get("last_modified") else None
        }
    }
    
    format_specific = {}
    if file_format == "parquet" and file_content:
        try:
            parquet_schema = extract_parquet_schema(file_content)
            format_specific["parquet"] = {
                "row_groups": parquet_schema.get("num_rows", 0),
                "compression": "unknown",
                "schema_version": "1.0"
            }
        except Exception as e:
            logger.warning('FN:extract_file_metadata file_name:{} file_format:{} error:{}'.format(file_name, file_format, str(e)))
    elif file_format == "csv" or file_format == "tsv":
        format_specific["csv"] = {
            "delimiter": "," if file_format == "csv" else "\t",
            "has_header": True,
            "encoding": "utf-8"
        }
    elif file_format == "json":
        format_specific["json"] = {
            "format": "unknown"
        }
    elif file_format == "avro" and file_content:
        format_specific["avro"] = {
            "format": "avro",
            "compression": "unknown"
        }
    elif file_format in ["xlsx", "xls"] and file_content:
        format_specific["excel"] = {
            "format": file_format,
            "has_header": True
        }
    elif file_format == "xml" and file_content:
        format_specific["xml"] = {
            "format": "xml",
            "encoding": "utf-8"
        }
    elif file_format == "orc" and file_content:
        format_specific["orc"] = {
            "format": "orc",
            "compression": "unknown"
        }
    elif file_format == "delta_lake" or "delta" in file_name.lower():
        format_specific["delta_lake"] = {
            "format": "delta_lake",
            "table_format": "delta"
        }
    elif is_data_lake:
        format_specific["data_lake"] = {
            "is_hdfs": True,
            "path_structure": "hierarchical"
        }
    
    if format_specific:
        file_metadata["format_specific"] = format_specific
    
    schema_json = None
    schema_hash = None
    

    if file_format == "parquet" and file_content:
        try:
            # Extract schema with PII detection (regex-based, no Azure DLP)
            # Uses column names for PII detection (no data samples needed)
            schema_json = extract_parquet_schema(file_content, include_pii_detection=True, sample_data=None)
            schema_hash = generate_schema_hash(schema_json)
        except Exception as e:
            logger.warning('FN:extract_file_metadata file_name:{} file_format:{} error:{}'.format(file_name, file_format, str(e)))
            schema_json = {}
            schema_hash = hashlib.shake_128(b"").hexdigest(16)
    elif file_format in ["csv", "tsv"] and file_content:
        try:
            schema_json = extract_csv_schema(file_content)
            schema_hash = generate_schema_hash(schema_json)
        except Exception as e:
            logger.warning('FN:extract_file_metadata file_name:{} file_format:{} error:{}'.format(file_name, file_format, str(e)))
            schema_json = {}
            schema_hash = hashlib.shake_128(b"").hexdigest(16)
    elif file_format == "json" and file_content:
        try:
            schema_json = extract_json_schema(file_content)
            schema_hash = generate_schema_hash(schema_json)
        except Exception as e:
            logger.warning('FN:extract_file_metadata file_name:{} file_format:{} error:{}'.format(file_name, file_format, str(e)))
            schema_json = {}
            schema_hash = hashlib.shake_128(b"").hexdigest(16)
    elif file_format == "avro" and file_content:
        try:
            schema_json = extract_avro_schema(file_content)
            schema_hash = generate_schema_hash(schema_json)
        except Exception as e:
            logger.warning('FN:extract_file_metadata file_name:{} file_format:{} error:{}'.format(file_name, file_format, str(e)))
            schema_json = {}
            schema_hash = hashlib.shake_128(b"").hexdigest(16)
    elif file_format in ["xlsx", "xls"] and file_content:
        try:
            schema_json = extract_excel_schema(file_content)
            schema_hash = generate_schema_hash(schema_json)
        except Exception as e:
            logger.warning('FN:extract_file_metadata file_name:{} file_format:{} error:{}'.format(file_name, file_format, str(e)))
            schema_json = {}
            schema_hash = hashlib.shake_128(b"").hexdigest(16)
    elif file_format == "xml" and file_content:
        try:
            schema_json = extract_xml_schema(file_content)
            schema_hash = generate_schema_hash(schema_json)
        except Exception as e:
            logger.warning('FN:extract_file_metadata file_name:{} file_format:{} error:{}'.format(file_name, file_format, str(e)))
            schema_json = {}
            schema_hash = hashlib.shake_128(b"").hexdigest(16)
    elif file_format == "orc" and file_content:
        try:
            schema_json = extract_orc_schema(file_content)
            schema_hash = generate_schema_hash(schema_json)
        except Exception as e:
            logger.warning('FN:extract_file_metadata file_name:{} file_format:{} error:{}'.format(file_name, file_format, str(e)))
            schema_json = {}
            schema_hash = hashlib.shake_128(b"").hexdigest(16)
    elif file_format == "delta_lake" or "delta" in file_name.lower():
        try:
            schema_json = extract_delta_lake_schema(file_content or b"", file_name)
            schema_hash = generate_schema_hash(schema_json)
        except Exception as e:
            logger.warning('FN:extract_file_metadata file_name:{} file_format:{} error:{}'.format(file_name, file_format, str(e)))
            schema_json = {}
            schema_hash = hashlib.shake_128(b"").hexdigest(16)
    else:

        schema_json = {
            "columns": [],
            "num_columns": 0,
            "format": file_format,
            "note": "Format not fully supported, basic metadata only"
        }
        schema_hash = hashlib.shake_128(json.dumps(schema_json).encode()).hexdigest(16)
    
    storage_metadata = {
        "azure": {
            "type": blob_info.get("blob_type", "Block blob"),
            "etag": blob_info.get("etag", "").strip('"') if blob_info.get("etag") else None,
            "access_tier": blob_info.get("access_tier"),
            "creation_time": blob_info.get("created_at").isoformat() if blob_info.get("created_at") else None,
            "last_modified": blob_info.get("last_modified").isoformat() if blob_info.get("last_modified") else None,
            "lease_status": blob_info.get("lease_status"),
            "content_encoding": blob_info.get("content_encoding"),
            "content_language": blob_info.get("content_language"),
            "cache_control": blob_info.get("cache_control"),
            "metadata": blob_info.get("metadata", {})
        }
    }
    
    return {
        "file_metadata": file_metadata,
        "schema_json": schema_json,
        "schema_hash": schema_hash,
        "file_hash": file_hash,
        "storage_metadata": storage_metadata
    }
